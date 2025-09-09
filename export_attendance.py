import pandas as pd
from datetime import datetime, timedelta
from babel.dates import format_date
import shutil
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from dateutil.parser import parse  # Add this import

def export_attendance(records):

    # Convert to list of dicts
    data = []
    for record in records:

        item = {}
        # Get username safely
        item["username"] = (
            record.get("user", {}).get("username", "N/A")
        )
        # Emplacement
        item["emplacement"] = record.get("emplacement", "N/A")
        # Format timestamp
        ts = parse(record["timestamp"])
        item["formatted_date"] = format_date(ts, format="d MMMM y", locale="mg_MG")

        data.append(item)

    # Create DataFrame
    df = pd.DataFrame(data)
    df = df[['username', 'emplacement', 'formatted_date']]
    df.columns = ['Anarana', 'Toerana', 'Daty']
    df = df.sort_values(by='Daty')

    # Export to Excel with styling
    export_path = "tmi_presence.xlsx"
    with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='tmi_presence')
        sheet = writer.sheets['tmi_presence']

        # Define borders and styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")

        # Style header row
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Style data cells and auto-width
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    cell.border = thin_border
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max_length + 2