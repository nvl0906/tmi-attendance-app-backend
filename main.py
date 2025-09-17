
from fastapi import FastAPI, UploadFile, File, Form, WebSocket, Body
from fastapi.middleware.cors import CORSMiddleware
import cv2
import numpy as np
import json
import base64
import io
from PIL import Image
from supabase import create_client, Client
from dotenv import load_dotenv
from insightface.app import FaceAnalysis
from datetime import datetime, timedelta, timezone
import os
import shutil
import asyncio
import pickle  # For saving/loading embeddings
from openpyxl import load_workbook
from fastapi.responses import FileResponse, JSONResponse
from scipy.spatial.distance import cosine
from liveness import liveness_check  # Import liveness_check function
from export_attendance import export_attendance # Import export_attendance to ensure it runs at startup
from merge_attendance import merge_attendance # Import merge_attendance to ensure it runs at startup    
import math

# Load .env
load_dotenv()

# Initialize Supabase
url: str = os.environ.get("SUPABASE_URL")
key: str = os.environ.get("SUPABASE_KEY")
supabase: Client = create_client(url, key)

# Memory cache
today_records_attendance = []
already_marked_today = set()

# Initialize FastAPI
app = FastAPI()

# CORS (optional for mobile app testing)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Change this in production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Load InsightFace model
face_app = FaceAnalysis(providers=['CUDAExecutionProvider'])
face_app.prepare(ctx_id=0, det_size=(640, 640))

known_faces = {}
PICKLE_FOLDER = "known_faces_embeddings"
os.makedirs(PICKLE_FOLDER, exist_ok=True)

def haversine(lat1, lon1, lat2, lon2):
    # Ensure all values are floats
    lat1, lon1, lat2, lon2 = map(float, [lat1, lon1, lat2, lon2])

    R = 6371000  # radius of Earth in meters
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)

    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    distance = R * c
    return round(distance)  # distance in meters

def get_member_id(username: str):
    response = supabase.table("members").select("*").eq("username", username).execute().data[0] if username else None
    if not response:
        raise ValueError("Member not found")
    return response["id"]

def get_emplacement():
    tz = timezone(timedelta(hours=3))
    now = datetime.now(tz)
    start_of_day = datetime.combine(now.date(), datetime.min.time(), tzinfo=tz)
    end_of_day = start_of_day + timedelta(days=1)

    response = supabase.table("attendance") \
        .select("emplacement, timestamp") \
        .gte("timestamp", start_of_day.isoformat()) \
        .lt("timestamp", end_of_day.isoformat()) \
        .limit(1) \
        .execute()

    data = response.data
    if data and len(data) > 0:
        value = data[0]["emplacement"]
    else:
        value = "aucun"
    return value


def load_today_attendance_one():
    global today_records_attendance
    tz = timezone(timedelta(hours=3))
    now = datetime.now(tz)
    start_of_day = datetime.combine(now.date(), datetime.min.time(), tzinfo=tz)
    end_of_day = start_of_day + timedelta(days=1)

    response = supabase.table("attendance")\
        .select("user:member_id(username), emplacement, timestamp")\
        .gte("timestamp", start_of_day.isoformat())\
        .lt("timestamp", end_of_day.isoformat())\
        .limit(1)\
        .execute().data 

    today_records_attendance = response if response else []

def load_today_attendance():
    global already_marked_today
    tz = timezone(timedelta(hours=3))
    now = datetime.now(tz)
    start_of_day = datetime.combine(now.date(), datetime.min.time(), tzinfo=tz)
    end_of_day = start_of_day + timedelta(days=1)

    response = supabase.table("attendance")\
        .select("user:member_id(username)")\
        .gte("timestamp", start_of_day.isoformat())\
        .lt("timestamp", end_of_day.isoformat())\
        .execute().data
    
    already_marked_today = {record["user"]["username"] for record in response} if response else set()

def get_users():
    response = supabase.table("members").select("username").execute().data
    return [user["username"] for user in response] if response else []

def normalize(vec):
    return vec / np.linalg.norm(vec)

def load_known_faces():
    known_faces.clear()  # clear previous data
    users = get_users()
    image_folder = "known_faces"

    for filename in os.listdir(image_folder):
        name, ext = os.path.splitext(filename)
        if name not in users:
            supabase.table("members").insert({"username": name, "voice": 1}).execute()
        if ext.lower() not in ['.jpg', '.png']:
            continue

        img_path = os.path.join(image_folder, filename)
        pkl_path = os.path.join(PICKLE_FOLDER, f"{name}.pkl")

        # Use pickle if it exists and is up-to-date
        if os.path.exists(pkl_path) and os.path.getmtime(pkl_path) >= os.path.getmtime(img_path):
            try:
                with open(pkl_path, "rb") as f:
                    emb = pickle.load(f)
                    known_faces[name] = emb
                    continue
            except Exception as e:
                print(f"⚠️ Error loading pickle for {name}: {e}")

        # If pickle not found or outdated, compute embedding from image
        img = cv2.imread(img_path)
        if img is None:
            continue
        img = cv2.copyMakeBorder(img, 50, 50, 50, 50, cv2.BORDER_CONSTANT, value=[255, 255, 255])
        try:
            faces = face_app.get(img)
            if faces:
                emb = normalize(faces[0].embedding)
                known_faces[name] = emb
                # Save embedding to pickle
                with open(pkl_path, "wb") as f:
                    pickle.dump(emb, f)
        except Exception as e:
            print(f"⚠️ Error processing image for {name}: {e}")

def match_face(embedding, threshold=0.59999):
    embedding = normalize(embedding)  # ensure input is normalized
    best_match = None
    best_distance = float("inf")

    for name, known_emb in known_faces.items():
        dist = cosine(embedding,known_emb)
        if dist < threshold and dist < best_distance:
            best_match = name
            best_distance = dist

    return best_match

@app.post("/register")
async def register_student(file: UploadFile, name: str = Form(...), voice: str = Form(...)):
    # Read image bytes
    contents = await file.read()
    np_img = np.frombuffer(contents, np.uint8)
    img = cv2.imdecode(np_img, cv2.IMREAD_COLOR)
    
    image_folder = "known_faces"
    for filename in os.listdir(image_folder):
        existingname, ext = os.path.splitext(filename)
        if existingname == name:
            return {"status": "username", "messageusername": f"{name} déjà pris!"}

    # Detect faces
    faces = face_app.get(img)
    if not faces:
        return {"status": "noface", "messagenoface": "Aucun visage détecté!"}

    # Check if exactly one face is detected
    if len(faces) != 1:
        return {"status": "lotface", "messagelotface": "Trop de visages détectés!"}

    # Perform liveness check
    if faces:
        f = faces[0]  # Use the first detected face
        x1, y1, x2, y2 = map(int, f.bbox.astype(int))
        bbox = (x1, y1, x2 - x1, y2 - y1)
        is_real, score = liveness_check(img, bbox, decision_threshold=0.400000)
        if not is_real:
            return {"status": "liveness", "messageliveness": "Veuillez vous bien centrer sinon fraude detecté!"}

    # Get embedding and check if it matches any known face
    new_embedding = normalize(faces[0].embedding)
    matched_name = match_face(new_embedding)

    if matched_name:
        return {"status": "existingface", "nameexisting": matched_name}

    save_path = f"./known_faces/{name}.jpg"
    with open(save_path, "wb") as buffer:
        buffer.write(contents)
    await asyncio.sleep(0.1)

    # Save embedding to pickle file
    emb_pkl_path = os.path.join(PICKLE_FOLDER, f"{name}.pkl")
    with open(emb_pkl_path, "wb") as f:
        pickle.dump(new_embedding, f)

    # Add embedding to memory
    known_faces[name] = new_embedding

    supabase.table("members").insert({"username": name, "voice": voice}).execute()

    return {"status": "success", "name": name}

@app.post("/recognize")
async def recognize(file: UploadFile = File(...), emplacement: str = Form(...), latitude: float = Form(...), longitude: float = Form(...)):
    tz = timezone(timedelta(hours=3))
    now = datetime.now(tz)
    start_of_day = datetime.combine(now.date(), datetime.min.time(), tzinfo=tz)
    end_of_day = start_of_day + timedelta(days=1)
    gps_data = supabase.table("gps")\
        .select("latitude,longitude")\
        .execute().data
    if not gps_data:
        return {"status": "noadmin", "messagenoadmin": "Aucun GPS activé par l'ADMIN aujourd'hui!"}
    res = gps_data[0]
    
    distance = haversine(res["latitude"], res["longitude"], latitude, longitude)

    if distance > 150:  # 100 meters threshold
        return {"status": "distance", "messagedistance": f"+ de 150m inacceptable: {distance}m!"} 
    # Read image bytes
    contents = await file.read()
    np_img = np.frombuffer(contents, np.uint8)
    img = cv2.imdecode(np_img, cv2.IMREAD_COLOR)

    # Detect faces
    faces = face_app.get(img)

    if not faces:
        return {"status": "noface", "messagenoface": "Aucun visage détecté!"}

    # Perform liveness check
    if len(faces) == 1:
        f = faces[0]  # Use the first detected face
        x1, y1, x2, y2 = map(int, f.bbox.astype(int))
        bbox = (x1, y1, x2 - x1, y2 - y1)
        is_real, score = liveness_check(img, bbox, decision_threshold=0.400000)
        if not is_real:
            return {"status": "liveness", "messageliveness": "Veuillez vous bien centrer sinon fraude detecté!"}

    matches = set()
    newly_marked = set()
    already_marked = set()
    timestamp = datetime.now(timezone(timedelta(hours=3)))
    start_of_day = datetime.combine(timestamp.date(), datetime.min.time(), tzinfo=timestamp.tzinfo)
    end_of_day = start_of_day + timedelta(days=1)

    if today_records_attendance:
        supabase_emplacement = today_records_attendance[0].get("emplacement", "").strip().lower()
        input_emplacement = emplacement.strip().lower()
        if supabase_emplacement != input_emplacement:
            return {"status": "emplacement_mismatch", "messageemplacement": f"Acceptable: {supabase_emplacement}"}

    # ✅ Continue recognition if emplacement is consistent or no records yet
    for i, face in enumerate(faces):
    
        emb = face.embedding
        name = match_face(emb)

        if name and name not in matches:
            matches.add(name)
            if name not in already_marked_today:
                 # Insert new attendance record
                supabase.table("attendance").insert({
                    "member_id": get_member_id(name),
                    "emplacement": emplacement
                }).execute()

                already_marked_today.add(name)
                newly_marked.add(name)
            elif name in already_marked_today:
                already_marked.add(name)
        else:
            pass
    load_today_attendance_one()
    if matches:
        return {"status": "success", "matches": list(matches), "newly_marked": list(newly_marked), "already_marked": list(already_marked)}
    else:
        return {"status": "nomatch", "messagenomatch": "Aucun visage reconnu!"}
    
@app.get("/download")
def download_excel():

    members = supabase.table("members").select("username, voice").execute().data

    # Path to Excel file
    list_path = "tmi_lisitra.xlsx"
    wb = load_workbook(list_path)
    ws = wb.active

    # Delete all rows except the first (header)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
        wb.save(list_path)

    # Append new members if not already present
    existing_names = [str(cell.value) for cell in ws['B'] if cell.value]
    for member in members:
        name = member.get("username")
        voice = member.get("voice")
        if name not in existing_names:
            ws.append([voice, name])
            wb.save(list_path)

    # Fetch all attendance records from Supabase
    response = supabase.table("attendance")\
        .select("user:member_id(username),emplacement,timestamp") \
        .execute().data
    export_attendance(response)
    merge_attendance()
    file_path = "tmi_presence_tracker_pivot_with_emplacement_split.xlsx"  # Or tmi_attendance.xlsx
    if os.path.exists(file_path):
        return FileResponse(
            path=file_path,
            filename="tmi_presence.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    return {"error": "Fichier non trouvé"}

@app.post("/login")
async def login(file: UploadFile = File(...), latitude: float = Form(...), longitude: float = Form(...)):
    load_today_attendance()
    load_today_attendance_one()
    # Read image bytes
    contents = await file.read()
    np_img = np.frombuffer(contents, np.uint8)
    img = cv2.imdecode(np_img, cv2.IMREAD_COLOR)
    
    # Detect faces
    faces = face_app.get(img)

    if not faces:
        return {"status": "noface", "messagenoface": "Aucun visage détecté!"}

    # Check if exactly one face is detected
    if len(faces) != 1:
        return {"status": "lotface", "messagelotface": "Trop de visages détectés!"}
    
    # Perform liveness check
    if len(faces) == 1:
        f = faces[0]  # Use the first detected face
        x1, y1, x2, y2 = map(int, f.bbox.astype(int))
        bbox = (x1, y1, x2 - x1, y2 - y1)
        is_real, score = liveness_check(img, bbox, decision_threshold=0.400000)
        if not is_real:
            return {"status": "liveness", "messageliveness": "Veuillez vous bien centrer sinon fraude detecté!"}

    # Get embedding and check if it matches any known face
    emb = faces[0].embedding
    name = match_face(emb)
    response = supabase.table("members").select("username,is_admin").eq("username", name).execute().data[0] if name else None
    if not response:
        return {"status": "notfound", "messagenotfound": "Veuillez vous enregistrer auprès de l'admin!"} 
    GPS_ID = "00000000-0000-0000-0000-000000000001"
    supabase_emplacement = get_emplacement()
    if response.get("is_admin") == True:
        supabase.table("gps").upsert({"id": GPS_ID, "latitude": latitude, "longitude": longitude}).execute()
        return {"status": "successadmin", "messageadmin": f"Bienvenue ADMIN {name}!", "sup_emplacement": supabase_emplacement}
    if response.get("is_admin") == False:
        return {"status": "successmember", "messagemember": f"Bienvenue {name}!", "sup_emplacement": supabase_emplacement}

@app.websocket("/ws/recognize")
async def ws_recognize(ws: WebSocket):
    await ws.accept()

    try:
        while True:
            message = await ws.receive_text()
            data = json.loads(message)  # Parse JSON from frontend
             # Extract base64 image and emplacement
            image_b64 = data.get("image")
            location =data.get("emplacement")
            emplacement = data.get("emplacement", "").strip().lower()
            try:
                if today_records_attendance:
                    supabase_emplacement = today_records_attendance[0].get("emplacement", "").strip().lower()
                    if supabase_emplacement != emplacement:
                        await ws.send_json({
                            "status": "emplacement_mismatch",
                            "messageemplacement": f"Acceptable: {supabase_emplacement}"
                        })
                        # Optionally close the connection if you want to force client to fix emplacement
                        await ws.close()
                        break
                
                # Decode Base64 image
                img_data = base64.b64decode(image_b64)
                np_img = np.frombuffer(img_data, np.uint8)
                img = cv2.imdecode(np_img, cv2.IMREAD_COLOR)

                # Detect faces
                faces = face_app.get(img)

                matches = set()
                newly_marked_today_view = set()
                already_marked_today_view = set()

                tz = timezone(timedelta(hours=3))
                now = datetime.now(tz)

                for face in faces:
                    emb = face.embedding
                    name = match_face(emb)

                    if name:
                        matches.add(name)

                        # Write to supabase only if not marked today
                        if name not in already_marked_today:
                            # Insert new attendance record
                            supabase.table("attendance").insert({
                                "member_id": get_member_id(name),
                                "emplacement": location
                            }).execute()
                            already_marked_today.add(name)
                            newly_marked_today_view.add(name)
                    
                        elif name in already_marked_today:
                            already_marked_today_view.add(name)

                # Send matches to client
                await ws.send_json({"status":"success", "users": list(matches), "already_marked": list(already_marked_today_view), "newly_marked": list(newly_marked_today_view)})

            except Exception as e:
                print("[Error] Processing frame:", e)

    except Exception as e:
        print(f"[WS] Error: {e}")

    finally:
        await ws.close()

@app.post("/search-user")
async def search_user(payload: dict = Body(...)):
    name = payload.get("name", "").strip()
    if not name or len(name) < 2:
        return {"user": None}
    # Search for user by username (case-insensitive, partial match)
    response = supabase.table("members")\
        .select("id, username, is_admin, voice")\
        .ilike("username", f"%{name}%")\
        .execute().data
    if response:
        return response
    return 

@app.post("/delete-user")
async def delete_user(payload: dict = Body(...)):
    user_id = payload.get("id")
    user_name = payload.get("name")
    # Delete from Supabase
    supabase.table("members").delete().eq("id", user_id).execute()
    # Optionally, remove face image and embedding
    image_path = os.path.join("known_faces", f"{user_name}.jpg")
    emb_path = os.path.join(PICKLE_FOLDER, f"{user_name}.pkl")
    for path in [image_path, emb_path]:
        if os.path.exists(path):
            os.remove(path)

    load_known_faces()
    load_today_attendance()
    load_today_attendance_one()

    return {"status": "success", "message": f"{user_name} supprimé avec succès!"}

@app.post("/update-user")
async def update_user(payload: dict = Body(...)):
    user_id = payload.get("id")
    new_name = payload.get("name")
    new_voice = payload.get("voice")
    new_admin = payload.get("is_admin")
    res = supabase.table("members").select("*").eq("id", user_id).execute().data[0]

    if new_name == res["username"] and new_voice == res["voice"] and new_admin == res["is_admin"] :
        return {"status": "error", "message": "Aucun changement détecté!"}
    
    # Update in Supabase
    if new_name != res["username"]:
        if new_name in get_users():
            return {"status": "error", "message": f"{new_name} est déjà pris!"}
        # Rename face image and embedding files
        old_image = os.path.join("known_faces", f"{res['username']}.jpg")
        old_emb = os.path.join(PICKLE_FOLDER, f"{res['username']}.pkl")
        new_image = os.path.join("known_faces", f"{new_name}.jpg")
        new_emb = os.path.join(PICKLE_FOLDER, f"{new_name}.pkl")
        if os.path.exists(old_image):
            os.rename(old_image, new_image)
        if os.path.exists(old_emb):
            os.rename(old_emb, new_emb)

        supabase.table("members").update({"username": new_name}).eq("id", user_id).execute()

        # Update memory
        if new_name in known_faces:
            known_faces[new_name] = known_faces.pop(new_name)

        supabase.table("members").update({"username": new_name}).eq("id", user_id).execute()

    if new_voice != res["voice"]:
        supabase.table("members").update({"voice": new_voice}).eq("id", user_id).execute()

    if new_admin != res["is_admin"]:
        supabase.table("members").update({"is_admin": new_admin}).eq("id", user_id).execute()

    load_known_faces()
    load_today_attendance()
    load_today_attendance_one()

    return {"status": "success", "message": f"{new_name} mis à jour avec succès!"}

load_known_faces()
load_today_attendance()
load_today_attendance_one()
