import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
import shutil
import openpyxl
from copy import copy

def merge_attendance():
    # Load full list
    df_list = pd.read_excel('tmi_lisitra.xlsx')

    # Load attendance export
    df_attendance = pd.read_excel('tmi_presence.xlsx')

    # Merge - keep all names from df_list, match attendance
    merged_rows = []
    for _, row in df_list.iterrows():
        name = row['Anarana']
        feo = row['Feo']
        matches = df_attendance[df_attendance['Anarana'] == name]
        if not matches.empty:
            for _, match in matches.iterrows():
                merged_rows.append({
                    'Feo': feo,
                    'Anarana': name,
                    'Toerana': match['Toerana'],
                    'Daty': match['Daty']
                })
        else:
            merged_rows.append({
                'Feo': feo,
                'Anarana': name,
                'Toerana': '',
                'Daty': ''
            })

    # Create merged DataFrame and export to Excel
    df_merged = pd.DataFrame(merged_rows)
    output_path = 'tmi_presence_tracker.xlsx'
    df_merged.to_excel(output_path, index=False)

    # Add border formatting
    wb = load_workbook(output_path)
    ws = wb.active
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value:
                cell.border = border
    wb.save(output_path)

    # --- Pivot and restructure for presence/absence by date ---
    # Load the merged data
    df_merged = pd.read_excel(output_path)

    # Get unique dates
    dates = sorted(df_merged['Daty'].dropna().unique())

    # Prepare the new DataFrame
    # Build DataFrame from unique (Feo, Anarana) pairs
    unique_pairs = df_merged[['Feo', 'Anarana']].drop_duplicates()
    result = unique_pairs.reset_index(drop=True)

    # For each date, mark present/absent
    for date in dates:
        result[date] = result['Anarana'].apply(
            lambda name: 'P' if not df_merged[(df_merged['Anarana'] == name) & (df_merged['Daty'] == date)].empty else 'A'
        )

    # --- Add Fahatongavana column: sum of "P" per row ---
    presence_cols = dates  # columns with P/A
    total_dates = len(presence_cols)
    fahatongavana_col = f"Fahatongavana({total_dates})"
    result[fahatongavana_col] = result[presence_cols].apply(lambda row: sum(val == "P" for val in row), axis=1)

    # Export to new Excel file
    output_pivot = 'tmi_presence_tracker_pivot.xlsx'
    result.to_excel(output_pivot, index=False)

    # --- Add emplacement row above date headers --

    # Reload merged data for emplacement lookup
    df_merged = pd.read_excel(output_path)
    # Reload pivoted data
    wb = openpyxl.load_workbook(output_pivot)
    ws = wb.active

    # Prepare emplacement row: first cell blank, then emplacement for each date

    # Build and store the main emplacement row for reuse in split sheets
    main_emplacement_row = ['','']
    for date in dates:
        loc = df_merged[df_merged['Daty'] == date]['Toerana'].dropna()
        main_emplacement_row.append(loc.iloc[0] if not loc.empty else '')

    # Insert emplacement row above headers
    ws.insert_rows(1)
    for col, value in enumerate(main_emplacement_row, 1):
        ws.cell(row=1, column=col, value=value)

    # Save to new file
    output_with_emplacement = 'tmi_presence_tracker_pivot_with_emplacement.xlsx'
    wb.save(output_with_emplacement)

    # --- Add borders to all cells with values in the final output file ---
    wb = openpyxl.load_workbook(output_with_emplacement)
    ws = wb.active
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != '':
                cell.border = border

    # Define fill colors
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # for 'present'
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # for 'absent'

    # Define center alignment
    center_alignment = Alignment(horizontal='center', vertical='center')
    bold_font = Font(bold=True)

    # Iterate over columns to apply formatting
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)

        for i, cell in enumerate(col):
            if cell.value:
                value = str(cell.value).strip()
                max_length = max(max_length, len(value))
                if value.lower() == "p":
                    cell.fill = green_fill
                elif value.lower() == "a":
                    cell.fill = red_fill

            # ✅ Center text in the first row
            if i == 0:
                cell.alignment = center_alignment
                cell.font = bold_font

        # Auto-adjust column width with padding
        ws.column_dimensions[column_letter].width = max_length + 2

    # Save the final output with all formatting 
    wb.save(output_with_emplacement)

    # --- Split into separate sheets by Feo value ---
    # Reload the final file as DataFrame to split
    df_final = pd.read_excel(output_with_emplacement, header=1)

    # Define Feo to sheet name mapping
    feo_sheet_map = {
        1: "Feo Voalohany",
        2: "Feo Faharoa",
        3: "Feo Fahatelo",
        4: "Feo Fahaefatra"
    }

    # Create a new workbook for split sheets
    split_wb = openpyxl.Workbook()
    # Remove the default sheet
    split_wb.remove(split_wb.active)

    for feo_value, sheet_name in feo_sheet_map.items():
        # Filter rows for this Feo value
        df_sheet = df_final[df_final['Feo'] == feo_value]
        if df_sheet.empty:
            continue
        ws_sheet = split_wb.create_sheet(title=sheet_name)


        # Prepare emplacement row for this sheet using precomputed main_emplacement_row
        columns = list(df_sheet.columns)
        date_columns = columns[2:]
        emplacement_row = ['', '']
        for date_col in date_columns:
            idx = dates.index(date_col) if date_col in dates else None
            if idx is not None:
                emplacement_row.append(main_emplacement_row[idx+2])
            else:
                emplacement_row.append('')

        # Insert emplacement row at the top
        ws_sheet.insert_rows(1)
        for col, value in enumerate(emplacement_row, 1):
            cell = ws_sheet.cell(row=1, column=col, value=value)
            # Center emplacement value and add border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Write headers in row 2
        for c_idx, col_name in enumerate(df_sheet.columns, start=1):
            cell = ws_sheet.cell(row=2, column=c_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # Add border to header row
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Write data rows starting from row 3
        for r_idx, row in enumerate(df_sheet.itertuples(index=False), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws_sheet.cell(row=r_idx, column=c_idx, value=value)
                # Center all cell values
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # Add border
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                # Color present/absent
                if isinstance(value, str):
                    if value.lower() == 'p':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value.lower() == 'a':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Bold Feo, Anarana, emplacement row, and date headers
        # Feo and Anarana columns are 1 and 2, emplacement row is row 1, date headers are from col 3
        ws_sheet.cell(row=1, column=1).font = Font(bold=True)
        ws_sheet.cell(row=1, column=2).font = Font(bold=True)
        for col in range(3, len(columns)+1):
            ws_sheet.cell(row=1, column=col).font = Font(bold=True)
            ws_sheet.cell(row=2, column=col).font = Font(bold=True)

        # --- Add rapport row (percentage of present for each date) ---
        # --- Add rapport row (percentage of present for each date) ---
        date_col_count = len(columns) - 3  # exclude Feo, Anarana, Fahatongavana
        rapport_row = ["", ""]  # For Feo and Anarana columns
        for idx in range(date_col_count):
            col_idx = idx + 3  # date columns start at column 3 (1-based index)
            present_count = 0
            total_count = ws_sheet.max_row - 2  # Exclude header and emplacement rows
            for row in range(3, ws_sheet.max_row + 1):
                val = ws_sheet.cell(row=row, column=col_idx).value
                if isinstance(val, str) and val.lower() == "p":
                    present_count += 1
            percent = (present_count / total_count * 100) if total_count > 0 else 0
            rapport_row.append(f"{percent:.1f}%")
        
        # Add empty for Fahatongavana column
        rapport_row.append("")
        ws_sheet.append(rapport_row)

        # Style rapport row
        for col in range(1, ws_sheet.max_column + 1):
            cell = ws_sheet.cell(row=ws_sheet.max_row, column=col)
            cell.font = Font(bold=True, color="1E90FF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


        # Auto-adjust column width
        for col in ws_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_sheet.column_dimensions[column_letter].width = max_length + 2


        # --- Add filters and sort rows by Anarana (column 2), EXCLUDING rapport row ---
        start_row = 3
        end_row = ws_sheet.max_row - 1  # Exclude rapport row
        max_col = ws_sheet.max_column

        # Extract rows with both value and style per cell
        rows_to_sort = []
        for row_idx in range(start_row, end_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = ws_sheet.cell(row=row_idx, column=col_idx)
                row_data.append({
                    "value": cell.value,
                    "style": {
                        "font": copy(cell.font),
                        "fill": copy(cell.fill),
                        "alignment": copy(cell.alignment),
                        "border": copy(cell.border),
                        "number_format": cell.number_format,
                    }
                })
            rows_to_sort.append(row_data)

        # Sort rows by Anarana (column 2), case-insensitive
        rows_to_sort.sort(key=lambda row: str(row[1]["value"]).lower() if row[1]["value"] else "")

        # Write back sorted rows with styles
        for i, row_data in enumerate(rows_to_sort, start=start_row):
            for j, cell_data in enumerate(row_data, start=1):
                cell = ws_sheet.cell(row=i, column=j)
                cell.value = cell_data["value"]
                cell.font = cell_data["style"]["font"]
                cell.fill = cell_data["style"]["fill"]
                cell.alignment = cell_data["style"]["alignment"]
                cell.border = cell_data["style"]["border"]
                cell.number_format = cell_data["style"]["number_format"]

        # The rapport row remains at the end and is not sorted

        # Add filter for data rows only (exclude rapport row)
        last_col = ws_sheet.max_column
        last_col_letter = get_column_letter(last_col)
        filter_end_row = ws_sheet.max_row - 1  # Exclude rapport row
        ws_sheet.auto_filter.ref = f"A2:{last_col_letter}{filter_end_row}"


        # Freeze pane at C3
        ws_sheet.freeze_panes = "C3"

        # Hide column A
        ws_sheet.column_dimensions['A'].hidden = True

    # Save the split workbook
    split_output = "tmi_presence_tracker_pivot_with_emplacement_split.xlsx"
    split_wb.save(split_output)


